import os
from datetime import datetime
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPORTS_DIR = "reports"


# ─────────────────────────────────────────────────────────
#  HTML Report
# ─────────────────────────────────────────────────────────

def generate_html_report(results: List[Dict], output_dir: str = REPORTS_DIR) -> str:
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"test_report_{timestamp}.html")

    total = len(results)
    passed = sum(1 for r in results if r.get("result") == "PASS")
    failed = total - passed
    pass_rate = round(passed / total * 100, 1) if total else 0
    avg_ms = round(sum(r.get("response_time_ms", 0) for r in results) / total, 1) if total else 0

    rows_html = ""
    for r in results:
        status = r.get("result", "N/A")
        row_cls = "pass-row" if status == "PASS" else "fail-row"
        rows_html += f"""
        <tr class="{row_cls}">
          <td><span class="tc-id">{r['id']}</span></td>
          <td>{r['name']}</td>
          <td><span class="api-tag">{r['api_name']}</span></td>
          <td class="mono">{r['endpoint']}</td>
          <td><span class="method method-{r['method'].lower()}">{r['method']}</span></td>
          <td><span class="badge badge-{r['test_type']}">{r['test_type'].title()}</span></td>
          <td class="mono">{r['expected_status']}</td>
          <td class="mono">{r.get('actual_status', '—')}</td>
          <td class="time">{r.get('response_time_ms', '—')} ms</td>
          <td><span class="result result-{'pass' if status == 'PASS' else 'fail'}">{status}</span></td>
          <td class="reason">{r.get('failure_reason', '')}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>API Test Report</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;background:#f0f2f5;color:#222}}
  .header{{background:linear-gradient(135deg,#0f0c29,#302b63,#24243e);color:#fff;padding:28px 40px}}
  .header h1{{font-size:1.8rem;letter-spacing:.5px}}
  .header p{{margin-top:6px;opacity:.75;font-size:.88rem}}
  .container{{padding:28px 40px}}
  .cards{{display:grid;grid-template-columns:repeat(5,1fr);gap:16px;margin-bottom:28px}}
  .card{{background:#fff;border-radius:12px;padding:20px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
  .card .val{{font-size:2.2rem;font-weight:700;line-height:1}}
  .card .lbl{{margin-top:6px;font-size:.75rem;text-transform:uppercase;letter-spacing:1px;color:#777}}
  .c-total .val{{color:#1a1a2e}}.c-pass .val{{color:#27ae60}}
  .c-fail .val{{color:#e74c3c}}.c-rate .val{{color:#2980b9}}
  .c-time .val{{color:#8e44ad;font-size:1.6rem}}
  .prog{{background:#eee;border-radius:8px;height:6px;margin-top:8px;overflow:hidden}}
  .prog-fill{{height:100%;border-radius:8px;background:linear-gradient(90deg,#27ae60,#2ecc71)}}
  .panel{{background:#fff;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.08);overflow:hidden}}
  .panel-head{{padding:16px 24px;border-bottom:1px solid #eee;display:flex;justify-content:space-between;align-items:center}}
  .panel-head h2{{font-size:1rem;color:#1a1a2e}}.panel-head span{{color:#999;font-size:.82rem}}
  table{{width:100%;border-collapse:collapse}}
  th{{background:#1a1a2e;color:#fff;padding:11px 14px;text-align:left;font-size:.8rem;text-transform:uppercase;letter-spacing:.4px}}
  td{{padding:10px 14px;border-bottom:1px solid #f2f2f2;font-size:.85rem;vertical-align:middle}}
  .pass-row:hover td{{background:#f0fff4}}.fail-row td{{background:#fffafa}}
  .method{{padding:3px 9px;border-radius:4px;font-weight:700;font-size:.75rem;color:#fff}}
  .method-get{{background:#27ae60}}.method-post{{background:#2980b9}}
  .method-put{{background:#e67e22}}.method-patch{{background:#8e44ad}}.method-delete{{background:#e74c3c}}
  .badge{{padding:2px 8px;border-radius:4px;font-size:.72rem;font-weight:600}}
  .badge-positive{{background:#e8f5e9;color:#2e7d32}}.badge-negative{{background:#fff3e0;color:#e65100}}
  .api-tag{{background:#e3f2fd;color:#1565c0;padding:2px 8px;border-radius:4px;font-size:.72rem;font-weight:500}}
  .tc-id{{font-family:monospace;background:#f5f5f5;padding:2px 6px;border-radius:4px;font-size:.8rem}}
  .result{{padding:3px 12px;border-radius:20px;font-weight:700;font-size:.78rem}}
  .result-pass{{background:#d4edda;color:#155724}}.result-fail{{background:#f8d7da;color:#721c24}}
  .mono{{font-family:monospace;font-size:.82rem;color:#555}}
  .time{{color:#999;font-family:monospace;font-size:.82rem}}
  .reason{{color:#c0392b;font-size:.78rem;font-style:italic;max-width:220px}}
  .footer{{text-align:center;padding:20px;color:#aaa;font-size:.78rem;margin-top:24px}}
</style>
</head>
<body>
<div class="header">
  <h1>AI-Assisted API Test Report</h1>
  <p>Generated {datetime.now().strftime('%B %d, %Y at %H:%M:%S')} &nbsp;&middot;&nbsp; Python + pytest + requests</p>
</div>
<div class="container">
  <div class="cards">
    <div class="card c-total"><div class="val">{total}</div><div class="lbl">Total Tests</div></div>
    <div class="card c-pass"><div class="val">{passed}</div><div class="lbl">Passed</div></div>
    <div class="card c-fail"><div class="val">{failed}</div><div class="lbl">Failed</div></div>
    <div class="card c-rate">
      <div class="val">{pass_rate}%</div><div class="lbl">Pass Rate</div>
      <div class="prog"><div class="prog-fill" style="width:{pass_rate}%"></div></div>
    </div>
    <div class="card c-time"><div class="val">{avg_ms}</div><div class="lbl">Avg Response (ms)</div></div>
  </div>
  <div class="panel">
    <div class="panel-head">
      <h2>Test Execution Results</h2>
      <span>{total} test cases &nbsp;&middot;&nbsp; 3 APIs &nbsp;&middot;&nbsp; 7 endpoints</span>
    </div>
    <table>
      <thead><tr>
        <th>Test ID</th><th>Test Name</th><th>API</th><th>Endpoint</th>
        <th>Method</th><th>Type</th><th>Exp. Status</th><th>Act. Status</th>
        <th>Resp. Time</th><th>Result</th><th>Failure Reason</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>
<div class="footer">AI-Assisted API Test Generation Framework &mdash; Built with Python &middot; pytest &middot; requests</div>
</body>
</html>"""

    with open(filepath, "w", encoding="utf-8") as fh:
        fh.write(html)
    return filepath


# ─────────────────────────────────────────────────────────
#  Excel Report
# ─────────────────────────────────────────────────────────

def generate_excel_report(results: List[Dict], output_dir: str = REPORTS_DIR) -> str:
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"test_report_{timestamp}.xlsx")

    rows = [
        {
            "Test ID": r["id"],
            "Test Name": r["name"],
            "API": r["api_name"],
            "Endpoint": r["endpoint"],
            "Method": r["method"],
            "Type": r["test_type"].title(),
            "Expected Status": r["expected_status"],
            "Actual Status": r.get("actual_status", "N/A"),
            "Response Time (ms)": r.get("response_time_ms", "N/A"),
            "Result": r.get("result", "N/A"),
            "Failure Reason": r.get("failure_reason", ""),
        }
        for r in results
    ]

    df = pd.DataFrame(rows)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Test Results", index=False)
        ws = writer.sheets["Test Results"]

        # Style header row
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Find "Result" column index
        result_col = next(
            (i for i, c in enumerate(ws[1], 1) if c.value == "Result"), None
        )

        pass_fill = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")
        pass_font = Font(color="276221", bold=True)
        fail_font = Font(color="9C0006", bold=True)

        for row in ws.iter_rows(min_row=2):
            if result_col:
                val = row[result_col - 1].value
                fill = pass_fill if val == "PASS" else fail_fill
                for cell in row:
                    cell.fill = fill
                row[result_col - 1].font = pass_font if val == "PASS" else fail_font

        # Auto-fit column widths
        for col_cells in ws.columns:
            width = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
                width + 4, 55
            )

    return filepath
